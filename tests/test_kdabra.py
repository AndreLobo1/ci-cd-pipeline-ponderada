from __future__ import annotations

from openpyxl import Workbook, load_workbook

from core.kdabra import generate_kdabra_enderecar_sheet


def test_generate_kdabra_enderecar_sheet_keeps_one_order_per_location(tmp_path):
    path = tmp_path / "kdabra.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Plano_Enderecamento_Final"
    ws.append(["location_id", "product_code"])
    ws.append(["WH001001-R2-013-4A", "SKU1"])
    ws.append(["WH001001-R2-013-4A", "SKU2"])
    ws.append(["WH001001-R2-013-4B", "SKU3"])
    wb.save(path)

    result = generate_kdabra_enderecar_sheet(path)

    assert result["success"] is True

    out_wb = load_workbook(path, data_only=True)
    out_ws = out_wb["kdabra enderecar"]
    rows = list(out_ws.iter_rows(values_only=True))

    assert rows[0] == ("galpao", "rua", "estante", "escaninho", "ordem")
    assert rows[1:] == [
        ("WH001001", "R2", "013", "4A", 1),
        ("WH001001", "R2", "013", "4B", 2),
    ]
