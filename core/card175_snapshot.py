from __future__ import annotations

import io
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .gsheets_client import CREDENTIALS_DIR, GSheetsClient
from .utils import normalize_string, parse_number

# Import lazy para evitar ciclo caso gsheets_backend importe card175 no futuro
def _generate_slots_from_cadastro(sheet_id: str, master_sheet_id: str | None = None) -> dict:
    from .gsheets_backend import generate_slots_from_cadastro_gsheet
    return generate_slots_from_cadastro_gsheet(sheet_id, master_sheet_id=master_sheet_id)

CARD175_PLAN_SHEET = "Plano_Enderecamento_Final_card_175"
CARD175_BASE_SHEET = "Plano_Enderecamento_Final_card_175_base"
CARD175_CHANGELOG_SHEET = "Log_Alteracoes_card_175"
WORKING_PLAN_SHEET = "Plano_Enderecamento_Final"
MAP_SHEET_FALLBACK = "Mapa_Final_Escaninhos"
CARD175_CONTEXT_PATH = CREDENTIALS_DIR / "card175_context.json"
UNALLOCATED_ID = "nao-alocado"
PRANCHETA_ID = "prancheta"
UNALLOCATED_LABEL = "NÃO ALOCADO"
ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

CARD175_REQUIRED_PLAN_HEADERS = [
    "location_id",
    "galpao_id",
    "rua_num",
    "equipamento_num",
    "tipo_equipamento",
    "nivel",
    "escaninho_num_no_nivel",
    "capacidade_l",
    "tipo_equipamento_final",
    "product_code",
    "product_name",
    "quantidade",
    "curva",
    "grupo",
    "categoria_armazenagem",
    "vol_l_unitario",
    "vol_L_unitario",
    "venda_total",
    "nm_fabricante",
    "altura_cm",
    "peso_kg_unitario",
    "subcategoria",
    "is_pesado",
    "is_alto",
    "is_hot_zone",
    "is_nivel_alto",
    "is_nivel_inferior",
    "is_realocado",
    "location_id_atual",
    "slot_duplo",
    "produto_alocado_code",
    "grupo_alocado",
]


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _normalize_header(value: Any) -> str:
    text = normalize_string(value).strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _find_index(headers: list[str], *aliases: str) -> int:
    aliases_norm = {_normalize_header(alias) for alias in aliases}
    for idx, header in enumerate(headers):
        if _normalize_header(header) in aliases_norm:
            return idx
    return -1


def _append_rows_chunked(client: GSheetsClient, sheet_name: str, rows: list[list[Any]], chunk_size: int = 400) -> None:
    if not rows:
        return
    for i in range(0, len(rows), chunk_size):
        client.append_rows(sheet_name, rows[i : i + chunk_size])


def _resolve_plan_source_sheet(client: GSheetsClient) -> tuple[str, str]:
    names = client.list_sheet_names()
    if WORKING_PLAN_SHEET in names:
        return WORKING_PLAN_SHEET, "plano"
    for name in names:
        norm = _normalize_header(name)
        if "plano_enderecamento" in norm:
            return name, "plano"
    if MAP_SHEET_FALLBACK in names:
        return MAP_SHEET_FALLBACK, "mapa"
    for name in names:
        norm = _normalize_header(name)
        if "mapa_final_escaninhos" in norm:
            return name, "mapa"
    raise ValueError(
        "Não encontrei aba de plano de endereçamento nem Mapa_Final_Escaninhos na planilha informada. "
        "Abas encontradas: " + ", ".join(names)
    )


def _ensure_required_headers(headers: list[str]) -> list[str]:
    normalized = {_normalize_header(item) for item in headers if normalize_string(item)}
    out = list(headers)
    for required in CARD175_REQUIRED_PLAN_HEADERS:
        key = _normalize_header(required)
        if key and key not in normalized:
            out.append(required)
            normalized.add(key)
    return out


def _to_float_qty(value: Any) -> float:
    parsed = parse_number(value)
    if parsed is None:
        return 0.0
    return float(parsed)


def _to_int_if_whole(value: float) -> int | float:
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return round(value, 4)


def _string_variants(value: Any, include_numeric: bool = True) -> set[str]:
    raw = normalize_string(value).upper()
    out: set[str] = set()
    if raw:
        out.add(raw)
        out.add(raw.lstrip("0") or "0")
        if raw.startswith("R") and raw[1:].isdigit():
            out.add(raw[1:].lstrip("0") or "0")
    if include_numeric:
        numeric_like = bool(re.fullmatch(r"[A-Z]?\d+", raw))
        num = parse_number(value)
        if num is not None and (numeric_like or raw.isdigit()):
            int_num = int(round(num))
            out.add(str(int_num))
            out.add(f"{int_num:03d}")
            out.add(f"R{int_num}")
    return {item for item in out if item}


def _register_addr_mapping(
    addr_map: dict[str, str],
    location_id: str,
    galpao: Any,
    rua: Any,
    posicao: Any,
    escaninho: Any,
) -> None:
    loc = normalize_string(location_id).strip()
    gal = normalize_string(galpao).upper().strip()
    if not loc or not gal:
        return
    ruas = _string_variants(rua, include_numeric=True) or {normalize_string(rua).upper()}
    posicoes = _string_variants(posicao, include_numeric=True) or {normalize_string(posicao).upper()}
    escaninhos = _string_variants(escaninho, include_numeric=False) or {normalize_string(escaninho).upper()}
    for rua_val in ruas:
        for pos_val in posicoes:
            for esc_val in escaninhos:
                key = f"{gal}|{rua_val}|{pos_val}|{esc_val}"
                addr_map.setdefault(key, loc)


def _extract_location_parts(location_id: str) -> tuple[str, str, str, str] | None:
    text = normalize_string(location_id).upper()
    match = re.match(r"^([A-Z0-9_]+)-R(\d+)-(?:E)?(\d+)-(.+)$", text)
    if not match:
        return None
    galpao = match.group(1)
    rua = match.group(2)
    pos = match.group(3)
    esc = match.group(4)
    return galpao, rua, pos, esc


def _letters_to_index(value: str) -> int | None:
    text = normalize_string(value).upper()
    if not text or not re.fullmatch(r"[A-Z]+", text):
        return None
    total = 0
    for ch in text:
        total = total * 26 + (ord(ch) - 64)
    return total


def _parse_slot_suffix(value: Any) -> tuple[int | None, int | None, str]:
    text = normalize_string(value).upper()
    if not text:
        return None, None, ""
    match = re.match(r"^(\d+)([A-Z]+)$", text)
    if not match:
        return None, None, text
    level = int(match.group(1))
    slot_letters = match.group(2)
    slot_num = _letters_to_index(slot_letters)
    return level, slot_num, text


def _build_virtual_location_id(galpao: str, rua: Any, posicao: Any, escaninho: Any) -> str | None:
    gal = normalize_string(galpao).upper()
    rua_num = parse_number(rua)
    equip_num = parse_number(posicao)
    _level, _slot_num, esc_text = _parse_slot_suffix(escaninho)
    if not gal or rua_num is None or equip_num is None or not esc_text:
        return None
    return f"{gal}-R{int(round(rua_num))}-{int(round(equip_num)):03d}-{esc_text}"


def _build_virtual_template_row(
    headers: list[str],
    location_id: str,
    group: dict[str, Any],
    equipment_template: list[Any] | None,
) -> list[Any]:
    row = list(equipment_template[: len(headers)]) if equipment_template else [""] * len(headers)
    row.extend([""] * (len(headers) - len(row)))

    galpao, rua, posicao, escaninho = _extract_location_parts(location_id) or ("", "", "", "")
    nivel_num, escaninho_num, esc_text = _parse_slot_suffix(escaninho)

    for idx, header in enumerate(headers):
        key = _normalize_header(header)
        if key == "location_id":
            row[idx] = location_id
        elif key == "galpao_id":
            row[idx] = galpao
        elif key == "rua_num":
            row[idx] = int(rua) if rua.isdigit() else row[idx]
        elif key == "equipamento_num":
            row[idx] = int(posicao) if posicao.isdigit() else row[idx]
        elif key == "nivel":
            row[idx] = nivel_num if nivel_num is not None else row[idx]
        elif key == "escaninho_num_no_nivel":
            row[idx] = escaninho_num if escaninho_num is not None else row[idx]
        elif key in {"product_code", "produto_alocado_code"}:
            row[idx] = "Vazio"
        elif key == "product_name":
            row[idx] = ""
        elif key == "quantidade":
            row[idx] = 0
        elif key == "slot_duplo":
            row[idx] = "NAO"
        elif key == "location_id_atual":
            row[idx] = location_id
        elif key == "grupo_alocado":
            row[idx] = ""

    if not equipment_template:
        for idx, header in enumerate(headers):
            key = _normalize_header(header)
            if key == "tipo_equipamento":
                row[idx] = "desconhecido"
            elif key == "tipo_equipamento_final":
                row[idx] = "desconhecido"
            elif key == "capacidade_l":
                row[idx] = 0
            elif key in {"is_hot_zone", "is_nivel_alto", "is_nivel_inferior", "is_realocado", "is_pesado", "is_alto"}:
                row[idx] = False

    return row


def _load_card175_upload_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb[wb.sheetnames[0]]
    values = list(sheet.iter_rows(min_row=1, values_only=True))
    if not values:
        return []
    headers = [normalize_string(col) for col in values[0]]
    rows = [
        {
            headers[idx]: raw[idx] if idx < len(raw) else None
            for idx in range(len(headers))
        }
        for raw in values[1:]
    ]
    return _normalize_card175_rows(rows)


def _normalize_card175_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not raw_rows:
        return []

    aggregated: dict[tuple[str, str, str, str, str, str, str], dict[str, Any]] = {}
    found_code_column = False
    for raw in raw_rows:
        if not isinstance(raw, dict):
            continue
        row_norm = {_normalize_header(k): v for k, v in raw.items()}
        code = normalize_string(
            row_norm.get("cod_produto")
            or row_norm.get("codigo_produto")
            or row_norm.get("sku")
            or row_norm.get("product_code")
        )
        if code:
            found_code_column = True
        else:
            continue
        id_local = normalize_string(row_norm.get("id_localizacao"))
        galpao = normalize_string(row_norm.get("galpao"))
        rua = normalize_string(row_norm.get("rua"))
        pos = normalize_string(
            row_norm.get("posicao_pallete")
            or row_norm.get("posicao_pallet")
            or row_norm.get("posicao")
            or row_norm.get("estante")
        )
        esc = normalize_string(
            row_norm.get("escaninho_nivel")
            or row_norm.get("escaninho")
            or row_norm.get("nivel")
            or row_norm.get("posicao_nivel")
        )
        desc = normalize_string(
            row_norm.get("desc_produto")
            or row_norm.get("descricao_produto")
            or row_norm.get("product_name")
            or row_norm.get("descricao")
        )
        qty = _to_float_qty(row_norm.get("quantidade") or row_norm.get("qtd") or row_norm.get("estoque"))
        key = (id_local, galpao, rua, pos, esc, code, desc)
        if key not in aggregated:
            aggregated[key] = {
                "id_localizacao": id_local,
                "galpao": galpao,
                "rua": rua,
                "posicao_pallete": pos,
                "escaninho_nivel": esc,
                "cod_produto": code,
                "desc_produto": desc,
                "quantidade": 0.0,
            }
        aggregated[key]["quantidade"] += qty
    if not found_code_column:
        raise ValueError("Card 175 não possui coluna de código do produto (cod_produto).")
    return list(aggregated.values())


def _build_external_location_maps(client: GSheetsClient) -> tuple[dict[str, str], dict[str, str]]:
    id_map: dict[str, str] = {}
    addr_map: dict[str, str] = {}
    for sheet_name in client.list_sheet_names():
        values = client.read_values(sheet_name)
        if not values or len(values) < 2:
            continue
        headers = [normalize_string(h) for h in values[0]]
        header_norm = [_normalize_header(h) for h in headers]
        idx_loc = _find_index(header_norm, "location_id", "location")
        if idx_loc == -1:
            continue
        idx_id_local = _find_index(header_norm, "id_localizacao")
        idx_galpao = _find_index(header_norm, "galpao", "galpao_id")
        idx_rua = _find_index(header_norm, "rua", "rua_num")
        idx_pos = _find_index(
            header_norm,
            "posicao_pallete",
            "posicao_pallet",
            "posicao",
            "equipamento_num",
            "equipamento",
            "estante",
        )
        idx_esc_composto = _find_index(header_norm, "escaninho_nivel")
        idx_nivel = _find_index(header_norm, "nivel")
        idx_esc_no_nivel = _find_index(header_norm, "escaninho_num_no_nivel", "escaninho")

        for raw in values[1:]:
            row = list(raw)
            location_id = normalize_string(row[idx_loc] if idx_loc < len(row) else "")
            if not location_id:
                continue
            if idx_id_local >= 0 and idx_id_local < len(row):
                id_local = normalize_string(row[idx_id_local])
                if id_local:
                    id_map.setdefault(id_local, location_id)
            if idx_galpao >= 0 and idx_rua >= 0 and idx_pos >= 0:
                gal = row[idx_galpao] if idx_galpao < len(row) else ""
                rua = row[idx_rua] if idx_rua < len(row) else ""
                pos = row[idx_pos] if idx_pos < len(row) else ""
                esc = ""
                if idx_esc_composto >= 0 and idx_esc_composto < len(row):
                    esc = normalize_string(row[idx_esc_composto])
                if not esc and idx_nivel >= 0 and idx_nivel < len(row) and idx_esc_no_nivel >= 0 and idx_esc_no_nivel < len(row):
                    nivel = normalize_string(row[idx_nivel])
                    esc_no_nivel = normalize_string(row[idx_esc_no_nivel])
                    if nivel and esc_no_nivel:
                        esc = f"{nivel}{esc_no_nivel}"
                if not esc and idx_esc_no_nivel >= 0 and idx_esc_no_nivel < len(row):
                    esc = normalize_string(row[idx_esc_no_nivel])
                if not esc and idx_nivel >= 0 and idx_nivel < len(row):
                    esc = normalize_string(row[idx_nivel])
                _register_addr_mapping(addr_map, location_id, gal, rua, pos, esc)
    return id_map, addr_map


def _build_base_products_map(client: GSheetsClient) -> dict[str, dict[str, Any]]:
    try:
        rows = client.read_sheet("Base_Produtos")
    except Exception:
        return {}
    mapping: dict[str, dict[str, Any]] = {}
    for row in rows:
        row_norm = {_normalize_header(k): v for k, v in row.items()}
        code = normalize_string(
            row_norm.get("product_code")
            or row_norm.get("codigo")
            or row_norm.get("cod_produto")
            or row_norm.get("sku")
        )
        if not code:
            continue
        mapping[code] = row_norm
    return mapping


def _set_card175_context(payload: dict[str, Any]) -> None:
    _save_json(CARD175_CONTEXT_PATH, payload)


def get_card175_context(sheet_id: str | None = None) -> dict[str, Any] | None:
    payload = _load_json(CARD175_CONTEXT_PATH)
    if not isinstance(payload, dict):
        return None
    if not payload.get("enabled"):
        return None
    if sheet_id and normalize_string(payload.get("sheet_id")) != normalize_string(sheet_id):
        return None
    return payload


def append_card175_change_logs(sheet_id: str, moves: list[dict[str, Any]], user: str = "local") -> dict[str, Any]:
    context = get_card175_context(sheet_id)
    if not context:
        return {"success": True, "logged": 0}

    changes_sheet = normalize_string(context.get("changes_sheet")) or CARD175_CHANGELOG_SHEET
    source_file = normalize_string(context.get("source_file"))
    client = GSheetsClient(sheet_id)
    client.ensure_sheet(changes_sheet)

    headers = [
        "timestamp",
        "usuario",
        "product_code",
        "location_id_original",
        "location_id_atual",
        "origem_operacao",
        "arquivo_card_175",
    ]
    existing = client.read_values(changes_sheet)
    if not existing or [normalize_string(h) for h in existing[0]] != headers:
        client.clear_sheet(changes_sheet)
        client.append_rows(changes_sheet, [headers])

    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    rows_to_append: list[list[Any]] = []
    for move in moves:
        loc_old_raw = normalize_string(move.get("locAnteriorId"))
        loc_new_raw = normalize_string(move.get("locNovoId"))
        if not loc_new_raw or loc_new_raw == PRANCHETA_ID:
            continue
        loc_old = loc_old_raw.replace("bin-", "") if loc_old_raw else UNALLOCATED_LABEL
        loc_new = loc_new_raw.replace("bin-", "") if loc_new_raw else UNALLOCATED_LABEL
        if loc_old_raw == UNALLOCATED_ID:
            loc_old = UNALLOCATED_LABEL
        if loc_new_raw == UNALLOCATED_ID:
            loc_new = UNALLOCATED_LABEL
        if normalize_string(loc_old) == normalize_string(loc_new):
            continue
        rows_to_append.append(
            [
                timestamp,
                user,
                normalize_string(move.get("productCode")),
                loc_old,
                loc_new,
                "DASHBOARD",
                source_file,
            ]
        )
    if rows_to_append:
        _append_rows_chunked(client, changes_sheet, rows_to_append)
    return {"success": True, "logged": len(rows_to_append), "changes_sheet": changes_sheet}


def import_card175_snapshot(
    sheet_id: str,
    file_bytes: bytes,
    source_file_name: str,
    user: str = "local",
) -> dict[str, Any]:
    if not file_bytes:
        return {"success": False, "error": "Arquivo vazio."}
    upload_rows = _load_card175_upload_rows(file_bytes)
    if not upload_rows:
        return {"success": False, "error": "Arquivo do card sem linhas válidas."}
    return _import_card175_normalized_rows(sheet_id, upload_rows, source_file_name, user=user)


def _import_card175_normalized_rows(
    sheet_id: str,
    upload_rows: list[dict[str, Any]],
    source_name: str,
    user: str = "local",
    master_sheet_id: str | None = None,
) -> dict[str, Any]:
    client = GSheetsClient(sheet_id)

    # Se não houver plano de endereçamento (ou existir mas vazio), gera a partir do Cadastro_Equipamentos.
    # Volumetria_Equipamentos vem da ETL (master_sheet_id), igual ao fluxo "endereçar novo warehouse".
    def _needs_generation(values: list) -> bool:
        return not values or len(values) < 2

    try:
        source_plan_sheet, source_kind = _resolve_plan_source_sheet(client)
        source_values = client.read_values(source_plan_sheet)
    except ValueError:
        source_plan_sheet = WORKING_PLAN_SHEET
        source_values = []

    if _needs_generation(source_values) and source_plan_sheet == WORKING_PLAN_SHEET:
        gen = _generate_slots_from_cadastro(sheet_id, master_sheet_id=master_sheet_id)
        if not gen.get("success"):
            return {
                "success": False,
                "error": (
                    "Não foi possível gerar Plano_Enderecamento_Final a partir do Cadastro_Equipamentos: "
                    f"{gen.get('error', 'erro desconhecido')}"
                ),
            }
        client = GSheetsClient(sheet_id)
        source_plan_sheet, source_kind = _resolve_plan_source_sheet(client)
        source_values = client.read_values(source_plan_sheet)

    if not source_values or len(source_values) < 2:
        return {"success": False, "error": f"Aba {source_plan_sheet} vazia após tentativa de geração."}

    source_headers = [normalize_string(h) for h in source_values[0]]
    plan_headers = _ensure_required_headers(source_headers)
    if not any(plan_headers):
        return {"success": False, "error": f"Cabeçalho inválido em {source_plan_sheet}."}

    idx_loc = _find_index(plan_headers, "location_id")
    idx_code = _find_index(plan_headers, "product_code")
    if idx_code == -1:
        idx_code = _find_index(plan_headers, "produto_alocado_code")
    if idx_loc == -1 or idx_code == -1:
        return {"success": False, "error": "Plano de origem sem colunas location_id/product_code."}

    idx_slot_duplo = _find_index(plan_headers, "slot_duplo")
    idx_name = _find_index(plan_headers, "product_name", "desc_produto")
    idx_qty = _find_index(plan_headers, "quantidade")
    idx_loc_atual = _find_index(plan_headers, "location_id_atual")
    idx_prod_alocado = _find_index(plan_headers, "produto_alocado_code")
    if idx_prod_alocado == -1:
        idx_prod_alocado = _find_index(plan_headers, "product_code")
    idx_group_alocado = _find_index(plan_headers, "grupo_alocado")

    extra_cols = len(plan_headers) - len(source_headers)
    data_rows = []
    for row in source_values[1:]:
        row_list = list(row)
        padded = row_list + [None] * (len(source_headers) - len(row_list))
        if extra_cols > 0:
            padded.extend([""] * extra_cols)
        data_rows.append(padded)
    template_by_loc: dict[str, list[Any]] = {}
    template_by_equipment: dict[tuple[str, str, str], list[Any]] = {}
    ordered_locations: list[str] = []
    for row in data_rows:
        loc = normalize_string(row[idx_loc]) if idx_loc < len(row) else ""
        if not loc:
            continue
        if loc not in template_by_loc:
            template_row = list(row[: len(plan_headers)])
            template_by_loc[loc] = template_row
            ordered_locations.append(loc)
            parts = _extract_location_parts(loc)
            if parts:
                gal, rua, pos, _esc = parts
                template_by_equipment.setdefault((gal, rua, pos), template_row)

    if not template_by_loc:
        return {"success": False, "error": "Nenhum location_id válido no plano de origem."}

    id_map, addr_map = _build_external_location_maps(client)
    for loc in ordered_locations:
        parts = _extract_location_parts(loc)
        if not parts:
            continue
        gal, rua, pos, esc = parts
        _register_addr_mapping(addr_map, loc, gal, rua, pos, esc)
    mapped_ordered_locations = list(ordered_locations)

    base_products = _build_base_products_map(client)
    aggregated_by_loc: dict[str, list[dict[str, Any]]] = defaultdict(list)
    unresolved_groups: dict[str, dict[str, Any]] = {}
    invalid_rows: list[list[Any]] = []
    unresolved = 0
    skipped_not_in_mix = 0
    for item in upload_rows:
        id_local = normalize_string(item.get("id_localizacao"))
        code = normalize_string(item.get("cod_produto"))
        if not code:
            continue
        if code not in base_products:
            skipped_not_in_mix += 1
            continue
        resolved_loc = ""
        if id_local and id_local in id_map:
            resolved_loc = id_map[id_local]
        if not resolved_loc:
            gal = normalize_string(item.get("galpao")).upper()
            rua_variants = _string_variants(item.get("rua"), include_numeric=True) or {normalize_string(item.get("rua")).upper()}
            pos_variants = _string_variants(item.get("posicao_pallete"), include_numeric=True) or {normalize_string(item.get("posicao_pallete")).upper()}
            esc_variants = _string_variants(item.get("escaninho_nivel"), include_numeric=False) or {normalize_string(item.get("escaninho_nivel")).upper()}
            for rua in rua_variants:
                if resolved_loc:
                    break
                for pos in pos_variants:
                    if resolved_loc:
                        break
                    for esc in esc_variants:
                        key = f"{gal}|{rua}|{pos}|{esc}"
                        mapped = addr_map.get(key)
                        if mapped:
                            resolved_loc = mapped
                            break
        if not resolved_loc or resolved_loc not in template_by_loc:
            unresolved += 1
            gal_raw = normalize_string(item.get("galpao")).upper()
            rua_raw = normalize_string(item.get("rua")).upper()
            pos_raw = normalize_string(item.get("posicao_pallete")).upper()
            esc_raw = normalize_string(item.get("escaninho_nivel")).upper()
            group_key = f"{gal_raw}|{rua_raw}|{pos_raw}|{esc_raw}"
            group = unresolved_groups.get(group_key)
            if not group:
                group = {
                    "galpao": gal_raw or "SEM_GALPAO",
                    "rua": rua_raw,
                    "posicao": pos_raw,
                    "escaninho": esc_raw,
                    "id_localizacao": normalize_string(item.get("id_localizacao")),
                    "items": [],
                }
                unresolved_groups[group_key] = group
            group["items"].append(item)
            continue
        aggregated_by_loc[resolved_loc].append(item)

    plan_rows_out: list[list[Any]] = []
    location_slot_counts: dict[str, int] = {}
    product_columns = [
        "product_code",
        "produto_alocado_code",
        "product_name",
        "quantidade",
        "curva",
        "grupo",
        "grupo_alocado",
        "categoria_armazenagem",
        "vol_l_unitario",
        "vol_l_unitario",
        "venda_total",
        "nm_fabricante",
        "altura_cm",
        "peso_kg_unitario",
        "subcategoria",
        "is_pesado",
        "is_alto",
        "is_pequeno",
        "is_fragil",
        "degelo",
        "metodo",
        "location_id_atual",
    ]
    product_col_indices = [idx for idx, h in enumerate(plan_headers) if _normalize_header(h) in product_columns]

    def clear_product_columns(row_data: list[Any]) -> None:
        for idx in product_col_indices:
            if idx < len(row_data):
                row_data[idx] = ""

    def apply_product(row_data: list[Any], location_id: str, product_data: dict[str, Any]) -> None:
        clear_product_columns(row_data)
        code = normalize_string(product_data.get("cod_produto"))
        desc = normalize_string(product_data.get("desc_produto"))
        qty = _to_int_if_whole(_to_float_qty(product_data.get("quantidade")))
        base = base_products.get(code, {})
        if idx_code >= 0:
            row_data[idx_code] = code
        if idx_prod_alocado >= 0:
            row_data[idx_prod_alocado] = code
        if idx_name >= 0:
            row_data[idx_name] = normalize_string(base.get("product_name") or desc)
        if idx_qty >= 0:
            row_data[idx_qty] = qty
        if idx_loc_atual >= 0:
            row_data[idx_loc_atual] = location_id
        if idx_group_alocado >= 0:
            row_data[idx_group_alocado] = normalize_string(base.get("grupo"))

        for idx, header in enumerate(plan_headers):
            key = _normalize_header(header)
            if key in {"product_code", "produto_alocado_code", "product_name", "quantidade", "location_id_atual", "grupo_alocado"}:
                continue
            if key in base and normalize_string(base.get(key)) != "":
                row_data[idx] = base.get(key)

    def apply_empty(row_data: list[Any], location_id: str) -> None:
        clear_product_columns(row_data)
        if idx_code >= 0:
            row_data[idx_code] = "Vazio"
        if idx_prod_alocado >= 0:
            row_data[idx_prod_alocado] = "Vazio"
        if idx_name >= 0:
            row_data[idx_name] = ""
        if idx_qty >= 0:
            row_data[idx_qty] = 0
        if idx_loc_atual >= 0:
            row_data[idx_loc_atual] = location_id

    def build_plan_rows_for_locations(location_ids: list[str]) -> tuple[list[list[Any]], dict[str, int], int]:
        rows_out: list[list[Any]] = []
        slot_counts: dict[str, int] = {}
        overflow_total = 0
        for location_id in location_ids:
            template = template_by_loc[location_id]
            rows_here = sorted(
                aggregated_by_loc.get(location_id, []),
                key=lambda item: _to_float_qty(item.get("quantidade")),
                reverse=True,
            )
            if not rows_here:
                out_row = list(template)
                apply_empty(out_row, location_id)
                rows_out.append(out_row)
                slot_counts[location_id] = 0
                continue

            keep = rows_here[:2]
            overflow = rows_here[2:]
            if overflow:
                overflow_total += len(overflow)
                for item in overflow:
                    invalid_rows.append(
                        [
                            "MAIS_DE_2_PRODUTOS_NO_ESCANINHO",
                            normalize_string(item.get("id_localizacao")),
                            normalize_string(item.get("galpao")),
                            normalize_string(item.get("rua")),
                            normalize_string(item.get("posicao_pallete")),
                            normalize_string(item.get("escaninho_nivel")),
                            normalize_string(item.get("cod_produto")),
                            normalize_string(item.get("desc_produto")),
                            _to_int_if_whole(_to_float_qty(item.get("quantidade"))),
                            location_id,
                        ]
                    )

            for item in keep:
                out_row = list(template)
                apply_product(out_row, location_id, item)
                rows_out.append(out_row)
            slot_counts[location_id] = len(keep)
        return rows_out, slot_counts, overflow_total

    plan_rows_out, location_slot_counts, overflow_count = build_plan_rows_for_locations(mapped_ordered_locations)

    virtual_locations_count = 0
    virtual_rows_added = 0

    for group_key in sorted(unresolved_groups.keys()):
        group = unresolved_groups[group_key]
        group_items = sorted(
            group.get("items", []),
            key=lambda item: _to_float_qty(item.get("quantidade")),
            reverse=True,
        )
        if not group_items:
            continue
        virtual_location_id = _build_virtual_location_id(
            normalize_string(group.get("galpao")),
            group.get("rua"),
            group.get("posicao"),
            group.get("escaninho"),
        )
        if virtual_location_id:
            template = template_by_loc.get(virtual_location_id)
            if template is None:
                parts = _extract_location_parts(virtual_location_id)
                if parts:
                    gal, rua, pos, _esc = parts
                    equipment_template = template_by_equipment.get((gal, rua, pos))
                else:
                    equipment_template = None
                template = _build_virtual_template_row(plan_headers, virtual_location_id, group, equipment_template)
                template_by_loc[virtual_location_id] = template
                ordered_locations.append(virtual_location_id)
                virtual_rows_added += 1
            aggregated_by_loc[virtual_location_id].extend(group_items)
            virtual_locations_count += len(group_items)
            continue

        for item in group_items:
            invalid_rows.append(
                [
                    "ENDERECO_NAO_MAPEADO",
                    normalize_string(item.get("id_localizacao")) or normalize_string(group.get("id_localizacao")),
                    normalize_string(item.get("galpao")) or normalize_string(group.get("galpao")),
                    normalize_string(item.get("rua")) or normalize_string(group.get("rua")),
                    normalize_string(item.get("posicao_pallete")) or normalize_string(group.get("posicao")),
                    normalize_string(item.get("escaninho_nivel")) or normalize_string(group.get("escaninho")),
                    normalize_string(item.get("cod_produto")),
                    normalize_string(item.get("desc_produto")),
                    _to_int_if_whole(_to_float_qty(item.get("quantidade"))),
                    "",
                ]
            )
        virtual_locations_count += len(group_items)

    if idx_slot_duplo == -1:
        plan_headers.append("slot_duplo")
        idx_slot_duplo = len(plan_headers) - 1
        for i in range(len(plan_rows_out)):
            plan_rows_out[i] = list(plan_rows_out[i]) + [""]
        for key, template_row in list(template_by_loc.items()):
            template_by_loc[key] = list(template_row) + [""]

    def apply_slot_duplo_flags(rows: list[list[Any]]) -> None:
        for row in rows:
            if len(row) < len(plan_headers):
                row.extend([""] * (len(plan_headers) - len(row)))
            loc = normalize_string(row[idx_loc]) if idx_loc < len(row) else ""
            slot_count = location_slot_counts.get(loc, 0)
            row[idx_slot_duplo] = "SIM" if slot_count >= 2 else "NAO"

    apply_slot_duplo_flags(plan_rows_out)

    client.clear_sheet(CARD175_PLAN_SHEET)
    _append_rows_chunked(client, CARD175_PLAN_SHEET, [plan_headers] + plan_rows_out)

    working_rows_out = list(plan_rows_out)
    if virtual_rows_added > 0:
        virtual_location_ids = [loc for loc in ordered_locations if loc not in mapped_ordered_locations]
        virtual_rows_out, virtual_slot_counts, virtual_overflow = build_plan_rows_for_locations(virtual_location_ids)
        overflow_count += virtual_overflow
        location_slot_counts.update(virtual_slot_counts)
        apply_slot_duplo_flags(virtual_rows_out)
        working_rows_out.extend(virtual_rows_out)

    apply_slot_duplo_flags(working_rows_out)

    client.clear_sheet(WORKING_PLAN_SHEET)
    _append_rows_chunked(client, WORKING_PLAN_SHEET, [plan_headers] + working_rows_out)

    client.clear_sheet(CARD175_BASE_SHEET)
    _append_rows_chunked(client, CARD175_BASE_SHEET, [plan_headers] + plan_rows_out)

    changelog_headers = [
        "timestamp",
        "usuario",
        "product_code",
        "location_id_original",
        "location_id_atual",
        "origem_operacao",
        "arquivo_card_175",
    ]
    client.clear_sheet(CARD175_CHANGELOG_SHEET)
    _append_rows_chunked(client, CARD175_CHANGELOG_SHEET, [changelog_headers])

    _set_card175_context(
        {
            "enabled": True,
            "sheet_id": sheet_id,
            "plan_sheet": CARD175_PLAN_SHEET,
            "working_sheet": WORKING_PLAN_SHEET,
            "base_sheet": CARD175_BASE_SHEET,
            "changes_sheet": CARD175_CHANGELOG_SHEET,
            "source_file": normalize_string(source_name),
            "imported_at": datetime.now().isoformat(timespec="seconds"),
        }
    )

    return {
        "success": True,
        "source_plan_sheet": source_plan_sheet,
        "source_sheet_kind": source_kind,
        "generated_sheet": CARD175_PLAN_SHEET,
        "rows_written": len(plan_rows_out),
        "working_rows_written": len(working_rows_out),
        "locations_total": len(ordered_locations),
        "locations_with_data": sum(1 for value in location_slot_counts.values() if value > 0),
        "invalid_rows": len(invalid_rows),
        "unmapped_rows": unresolved,
        "overflow_rows": overflow_count,
        "skipped_not_in_mix": skipped_not_in_mix,
        "virtual_locations": virtual_locations_count,
        "virtual_address_groups": len(unresolved_groups),
        "virtual_rows_added": virtual_rows_added,
        "generated_sheet_url": client.get_sheet_url(CARD175_PLAN_SHEET),
        "working_sheet_url": client.get_sheet_url(WORKING_PLAN_SHEET),
        "changes_sheet_url": client.get_sheet_url(CARD175_CHANGELOG_SHEET),
    }


def import_card175_rows(
    sheet_id: str,
    rows: list[dict[str, Any]],
    source_name: str,
    user: str = "local",
    master_sheet_id: str | None = None,
) -> dict[str, Any]:
    normalized_rows = _normalize_card175_rows(rows)
    if not normalized_rows:
        return {"success": False, "error": "Card 175 sem linhas válidas."}
    return _import_card175_normalized_rows(sheet_id, normalized_rows, source_name, user=user, master_sheet_id=master_sheet_id)
