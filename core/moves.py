from __future__ import annotations

import threading
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

import re

from .utils import normalize_string, parse_bool_flag, parse_number

PRANCHETA_ID = "PRANCHETA"
UNALLOCATED_ID = "UNALLOCATED"
LOG_UNALLOCATED_LABEL = "NÃO ALOCADO"
LOG_DATE_FORMAT = "%d/%m/%Y"
LOG_TIME_FORMAT = "%H:%M:%S"

_lock = threading.Lock()


def _get_log_datetime() -> tuple[str, str]:
    now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    return now.strftime(LOG_DATE_FORMAT), now.strftime(LOG_TIME_FORMAT)


def _normalize_header(header: Any) -> str:
    return normalize_string(header).lower().replace(" ", "_")


def _ensure_log_columns(ws) -> list[str]:
    headers = [normalize_string(cell.value) for cell in ws[1]]
    normalized = [_normalize_header(h) for h in headers]
    if "product_name" not in normalized:
        if "product_code" in normalized:
            idx_product = normalized.index("product_code")
            ws.insert_cols(idx_product + 2)
            ws.cell(row=1, column=idx_product + 2, value="product_name")
            headers.insert(idx_product + 1, "product_name")
            normalized.insert(idx_product + 1, "product_name")
        else:
            ws.insert_cols(1)
            ws.cell(row=1, column=1, value="product_name")
            headers.insert(0, "product_name")
            normalized.insert(0, "product_name")

    idx_data = normalized.index("data") if "data" in normalized else -1
    idx_hora = normalized.index("hora") if "hora" in normalized else -1
    idx_data_mov = normalized.index("data_movimentacao") if "data_movimentacao" in normalized else -1

    if idx_data != -1 and idx_hora != -1:
        return headers

    if idx_data_mov != -1:
        headers[idx_data_mov] = "data"
        ws.cell(row=1, column=idx_data_mov + 1, value="data")
        if idx_hora == -1:
            ws.insert_cols(idx_data_mov + 2)
            ws.cell(row=1, column=idx_data_mov + 2, value="hora")
            headers.insert(idx_data_mov + 1, "hora")
        return headers

    # append at end
    ws.insert_cols(len(headers) + 1, amount=2)
    ws.cell(row=1, column=len(headers) + 1, value="data")
    ws.cell(row=1, column=len(headers) + 2, value="hora")
    headers.extend(["data", "hora"])
    return headers


def _build_log_row(headers: list[str], entry: dict[str, Any]) -> list[Any]:
    row = [""] * len(headers)
    data_hora = f"{entry.get('data','')} {entry.get('hora','')}".strip()
    for idx, header in enumerate(headers):
        key = _normalize_header(header)
        if key in {"product_code", "codigo_produto", "produto"}:
            row[idx] = entry.get("product_code") or ""
        elif key in {"product_name", "nome_produto", "produto_nome", "descricao_produto", "desc_produto"}:
            row[idx] = entry.get("product_name") or ""
        elif key in {"location_id_anterior", "loc_anterior", "origem"}:
            row[idx] = entry.get("location_id_anterior") or ""
        elif key in {"location_id_novo", "loc_novo", "destino"}:
            row[idx] = entry.get("location_id_novo") or ""
        elif key == "data":
            row[idx] = entry.get("data") or ""
        elif key == "hora":
            row[idx] = entry.get("hora") or ""
        elif key == "data_movimentacao":
            row[idx] = data_hora
        elif key in {"motivo", "acao"}:
            row[idx] = entry.get("motivo") or ""
        elif key in {"usuario", "user", "email"}:
            row[idx] = entry.get("usuario") or ""
    return row


def _build_new_row_value(header: str, index: int, product_info: dict[str, Any] | None, original_row: list[Any], headers: list[str]) -> Any:
    p = product_info or {}
    is_vazio = not product_info or normalize_string(p.get("product_code")) == "Vazio"
    header_key = normalize_string(header).lower()

    if header_key in {"product_code", "produto_alocado_code"}:
        return None if is_vazio else normalize_string(p.get("product_code")) or None
    if header_key == "product_name":
        return None if is_vazio else p.get("product_name")
    if header_key == "curva":
        return None if is_vazio else p.get("curva")
    if header_key in {"grupo", "grupo_alocado"}:
        return None if is_vazio else p.get("grupo")
    if header_key == "categoria_armazenagem":
        return None if is_vazio else p.get("categoria_armazenagem")
    if header_key == "vol_l_unitario":
        return None if is_vazio else (p.get("vol_l_unitario") or p.get("vol_L_unitario"))
    if header_key == "quantidade":
        return None if is_vazio else p.get("quantidade")
    if header_key == "venda_total":
        return None if is_vazio else p.get("venda_total")
    if header_key == "nm_fabricante":
        return None if is_vazio else p.get("nm_fabricante")
    if header_key == "altura_cm":
        return None if is_vazio else p.get("altura_cm")
    if header_key == "peso_kg_unitario":
        return None if is_vazio else p.get("peso_kg_unitario")
    if header_key == "subcategoria":
        return None if is_vazio else p.get("subcategoria")
    if header_key == "is_pesado":
        return False if is_vazio else parse_bool_flag(p.get("is_pesado"))
    if header_key == "is_alto":
        return False if is_vazio else parse_bool_flag(p.get("is_alto"))
    if header_key == "is_realocado":
        return not is_vazio
    if header_key == "location_id_atual":
        loc_idx = headers.index("location_id") if "location_id" in headers else -1
        return None if is_vazio else (original_row[loc_idx] if loc_idx != -1 and loc_idx < len(original_row) else None)

    return original_row[index] if index < len(original_row) else None


def _update_row(ws, row_num: int, headers: list[str], product_info: dict[str, Any] | None, original_row: list[Any]) -> None:
    for col_index, header in enumerate(headers, start=1):
        value = _build_new_row_value(header, col_index - 1, product_info, original_row, headers)
        cell = ws.cell(row=row_num, column=col_index)
        cell.value = value


def _append_logs(ws, entries: list[dict[str, Any]]) -> None:
    if not entries:
        return
    headers = _ensure_log_columns(ws)
    for entry in entries:
        ws.append(_build_log_row(headers, entry))


def _find_header_index(headers: list[str], *keys: str) -> int:
    for key in keys:
        target = _normalize_header(key)
        for idx, header in enumerate(headers):
            if _normalize_header(header) == target:
                return idx
    return -1


def _parse_equip_id(value: Any) -> tuple[int | None, int | None]:
    text = normalize_string(value).upper().replace(" ", "")
    if not text:
        return None, None
    match = re.match(r"R(\d+)(?:-?E?)(\d+)$", text)
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _extract_product_info_from_row(row_data: list[Any], headers: list[str]) -> dict[str, Any]:
    product_info: dict[str, Any] = {}

    product_columns = {
        "product_code",
        "produto_alocado_code",
        "product_name",
        "curva",
        "grupo",
        "grupo_alocado",
        "categoria_armazenagem",
        "vol_l_unitario",
        "vol_l_unitario",
        "quantidade",
        "venda_total",
        "nm_fabricante",
        "altura_cm",
        "peso_kg_unitario",
        "subcategoria",
        "is_pesado",
        "is_alto",
    }

    for index, header in enumerate(headers):
        key = _normalize_header(header)
        if key in product_columns:
            product_info[key] = row_data[index] if index < len(row_data) else None

    if not product_info.get("product_code") and product_info.get("produto_alocado_code"):
        product_info["product_code"] = product_info.get("produto_alocado_code")
    if not product_info.get("product_code"):
        product_info["product_code"] = "Vazio"

    return product_info


def _build_location_map(ws, headers: list[str]) -> dict[str, list[int]]:
    if "location_id" not in headers:
        return {}
    loc_idx = headers.index("location_id") + 1
    location_map: dict[str, list[int]] = {}
    for row in range(2, ws.max_row + 1):
        loc_value = ws.cell(row=row, column=loc_idx).value
        loc_id = normalize_string(loc_value)
        if loc_id:
            location_map.setdefault(loc_id, []).append(row)
    return location_map


def save_batch_moves(path: Path, moves: list[dict[str, Any]], user: str = "local") -> dict[str, Any]:
    with _lock:
        wb = load_workbook(path)
        plano_ws = wb["Plano_Enderecamento_Final"]
        log_ws = wb["Log_Reenderecamento"] if "Log_Reenderecamento" in wb.sheetnames else None

        headers = [normalize_string(cell.value) for cell in plano_ws[1]]
        slot_duplo_added = False
        if "slot_duplo" not in headers:
            headers.append("slot_duplo")
            plano_ws.cell(row=1, column=len(headers), value="slot_duplo")
            slot_duplo_added = True
        location_map = _build_location_map(plano_ws, headers)
        product_idx = _find_header_index(headers, "product_code", "produto_alocado_code")
        if product_idx == -1:
            return {"success": False, "error": "Coluna product_code não encontrada no Plano_Enderecamento_Final."}
        slot_duplo_idx = _find_header_index(headers, "slot_duplo")

        original_last_row = plano_ws.max_row
        row_states: dict[int, list[Any]] = {}
        for row_num in range(2, plano_ws.max_row + 1):
            row_values = [cell.value for cell in plano_ws[row_num]]
            if len(row_values) < len(headers):
                row_values.extend([None] * (len(headers) - len(row_values)))
            row_states[row_num] = row_values[: len(headers)]
        modified_rows: set[int] = set()
        affected_locations: set[str] = set()

        logs: list[dict[str, Any]] = []
        missing_targets: list[str] = []
        full_targets: list[str] = []
        data, hora = _get_log_datetime()
        prepared_moves: list[dict[str, Any]] = []

        def _row_values(row_num: int) -> list[Any]:
            row = row_states.get(row_num)
            if row is None:
                return [None] * len(headers)
            return list(row)

        def _row_code(row_num: int) -> str:
            values = _row_values(row_num)
            return normalize_string(values[product_idx] if product_idx < len(values) else None)

        def _is_row_empty(row_num: int) -> bool:
            code = _row_code(row_num)
            return not code or code == "Vazio"

        def _write_row(row_num: int, product_info: dict[str, Any] | None) -> None:
            original_row = _row_values(row_num)
            row_states[row_num] = [
                _build_new_row_value(header, idx, product_info, original_row, headers)
                for idx, header in enumerate(headers)
            ]
            modified_rows.add(row_num)

        def _find_source_row(location_id: str | None, product_code: str) -> int | None:
            if not location_id:
                return None
            loc_rows = location_map.get(location_id) or []
            if not loc_rows:
                return None
            for row_num in loc_rows:
                if _row_code(row_num) == product_code:
                    return row_num
            for row_num in loc_rows:
                if not _is_row_empty(row_num):
                    return row_num
            return None

        def _clone_row_for_location(location_id: str) -> int | None:
            loc_rows = location_map.get(location_id) or []
            if not loc_rows:
                return None
            base_row_num = loc_rows[0]
            base_values = _row_values(base_row_num)
            new_row_num = (max(row_states.keys()) + 1) if row_states else 2
            row_states[new_row_num] = base_values
            location_map.setdefault(location_id, []).append(new_row_num)
            return new_row_num

        def _find_dest_row(location_id: str | None) -> tuple[int | None, str | None]:
            if not location_id:
                return None, "missing"
            loc_rows = location_map.get(location_id) or []
            if not loc_rows:
                return None, "missing"
            for row_num in loc_rows:
                if _is_row_empty(row_num):
                    return row_num, None
            occupied_count = sum(0 if _is_row_empty(rn) else 1 for rn in loc_rows)
            if occupied_count >= 2:
                return None, "full"
            new_row = _clone_row_for_location(location_id)
            if not new_row:
                return None, "missing"
            return new_row, None

        for move in moves:
            product_code = normalize_string(move.get("productCode"))
            product_info = move.get("productInfo") or {}
            product_name = normalize_string(
                product_info.get("product_name")
                or product_info.get("nome")
                or move.get("productName")
            )
            loc_anterior = normalize_string(move.get("locAnteriorId"))
            loc_novo = normalize_string(move.get("locNovoId"))

            if loc_novo == PRANCHETA_ID:
                continue

            clean_anterior = loc_anterior.replace("bin-", "") if loc_anterior else None
            clean_novo = loc_novo.replace("bin-", "") if loc_novo else None

            if clean_anterior and loc_anterior not in {PRANCHETA_ID, UNALLOCATED_ID}:
                src_row_num = _find_source_row(clean_anterior, product_code)
                if src_row_num:
                    affected_locations.add(clean_anterior)
                    _write_row(src_row_num, None)

            prepared_moves.append(
                {
                    "product_code": product_code,
                    "product_info": product_info,
                    "product_name": product_name,
                    "loc_anterior": loc_anterior,
                    "loc_novo": loc_novo,
                    "clean_anterior": clean_anterior,
                    "clean_novo": clean_novo,
                    "expects_target": bool(loc_novo and loc_novo not in {UNALLOCATED_ID, PRANCHETA_ID}),
                }
            )

        for item in prepared_moves:
            product_code = item["product_code"]
            product_info = item["product_info"]
            product_name = item["product_name"]
            loc_anterior = item["loc_anterior"]
            loc_novo = item["loc_novo"]
            clean_anterior = item["clean_anterior"]
            clean_novo = item["clean_novo"]
            expects_target = item["expects_target"]

            if loc_novo == UNALLOCATED_ID:
                log_anterior_label = LOG_UNALLOCATED_LABEL if loc_anterior == UNALLOCATED_ID else clean_anterior
                log_novo_label = LOG_UNALLOCATED_LABEL
                if log_anterior_label:
                    logs.append(
                        {
                            "product_code": product_code,
                            "product_name": product_name,
                            "location_id_anterior": log_anterior_label,
                            "location_id_novo": log_novo_label,
                            "data": data,
                            "hora": hora,
                            "motivo": "MANUAL-REEND",
                            "usuario": user,
                        }
                    )
                continue

            dest_row = None
            if expects_target:
                dest_row, dest_err = _find_dest_row(clean_novo)
                if dest_err == "missing":
                    if clean_novo:
                        missing_targets.append(clean_novo)
                    continue
                if dest_err == "full":
                    if clean_novo:
                        full_targets.append(clean_novo)
                    continue
                if not dest_row:
                    if clean_novo:
                        missing_targets.append(clean_novo)
                    continue
                affected_locations.add(clean_novo)

            if dest_row:
                _write_row(dest_row, product_info)

            log_anterior_label = LOG_UNALLOCATED_LABEL if loc_anterior == UNALLOCATED_ID else clean_anterior
            log_novo_label = LOG_UNALLOCATED_LABEL if loc_novo == UNALLOCATED_ID else clean_novo
            if (log_anterior_label or log_novo_label) and loc_novo != PRANCHETA_ID:
                logs.append(
                    {
                        "product_code": product_code,
                        "product_name": product_name,
                        "location_id_anterior": log_anterior_label,
                        "location_id_novo": log_novo_label,
                        "data": data,
                        "hora": hora,
                        "motivo": "MANUAL-REEND",
                        "usuario": user,
                    }
                )

        if missing_targets:
            missing_unique = sorted(set(missing_targets))
            preview = ", ".join(missing_unique[:5])
            suffix = "..." if len(missing_unique) > 5 else ""
            return {
                "success": False,
                "error": f"Destino(s) não encontrado(s) no Plano_Enderecamento_Final: {preview}{suffix}",
                "missingTargets": missing_unique,
            }
        if full_targets:
            full_unique = sorted(set(full_targets))
            preview = ", ".join(full_unique[:5])
            suffix = "..." if len(full_unique) > 5 else ""
            return {
                "success": False,
                "error": f"Destino(s) cheio(s) (2 produtos por escaninho): {preview}{suffix}",
                "fullTargets": full_unique,
            }

        if slot_duplo_idx >= 0:
            slot_locations = set(location_map.keys()) if slot_duplo_added else affected_locations
            if not slot_locations:
                slot_locations = set(location_map.keys())
            for location_id in slot_locations:
                loc_rows = location_map.get(location_id) or []
                if not loc_rows:
                    continue
                occupied = sum(0 if _is_row_empty(rn) else 1 for rn in loc_rows)
                flag = "SIM" if occupied >= 2 else "NAO"
                for row_num in loc_rows:
                    row = row_states.get(row_num)
                    if not row:
                        continue
                    if slot_duplo_idx >= len(row):
                        row.extend([None] * (slot_duplo_idx - len(row) + 1))
                    if row[slot_duplo_idx] != flag:
                        row[slot_duplo_idx] = flag
                        modified_rows.add(row_num)

        updates = 0
        appended = 0
        for row_num in sorted(modified_rows):
            values = _row_values(row_num)[: len(headers)]
            if row_num <= original_last_row:
                for col_index, value in enumerate(values, start=1):
                    plano_ws.cell(row=row_num, column=col_index, value=value)
                updates += 1
            else:
                plano_ws.append(values)
                appended += 1

        if log_ws:
            _append_logs(log_ws, logs)

        wb.save(path)

    return {
        "success": True,
        "processed": len(moves),
        "updated": updates,
        "appended": appended,
        "logsAdded": len(logs),
    }


def save_single_move(path: Path, move: dict[str, Any], user: str = "local") -> dict[str, Any]:
    return save_batch_moves(path, [move], user=user)


def execute_swap(path: Path, swap_info: dict[str, Any], user: str = "local") -> dict[str, Any]:
    move_a = swap_info.get("moveA", {})
    move_b = swap_info.get("moveB", {})

    loc_a = normalize_string(move_a.get("locAnteriorId"))
    loc_b = normalize_string(move_b.get("locAnteriorId"))
    product_a = move_a.get("productInfo") or {}
    product_b = move_b.get("productInfo") or {}

    if not loc_a or not loc_b:
        return {"success": False, "error": "Locais de troca inválidos"}

    swap_moves = [
        {"productCode": normalize_string(move_b.get("productCode")), "locAnteriorId": loc_b, "locNovoId": loc_a, "productInfo": product_b},
        {"productCode": normalize_string(move_a.get("productCode")), "locAnteriorId": loc_a, "locNovoId": loc_b, "productInfo": product_a},
    ]

    return save_batch_moves(path, swap_moves, user=user)


def execute_equipment_swap(path: Path, equip_a_id: Any, equip_b_id: Any, user: str = "local") -> dict[str, Any]:
    with _lock:
        wb = load_workbook(path)
        if "Plano_Enderecamento_Final" not in wb.sheetnames:
            return {"success": False, "error": "Aba Plano_Enderecamento_Final não encontrada."}

        plano_ws = wb["Plano_Enderecamento_Final"]
        log_ws = wb["Log_Reenderecamento"] if "Log_Reenderecamento" in wb.sheetnames else None

        headers_raw = [normalize_string(cell.value) for cell in plano_ws[1]]
        headers = [_normalize_header(h) for h in headers_raw]

        rua_idx = _find_header_index(headers, "rua_num", "rua")
        equip_idx = _find_header_index(headers, "equipamento_num", "equipamento")
        tipo_idx = _find_header_index(headers, "tipo_equipamento", "tipo")

        if rua_idx == -1 or equip_idx == -1 or tipo_idx == -1:
            return {
                "success": False,
                "error": "Colunas essenciais (rua_num, equipamento_num, tipo_equipamento) não encontradas.",
            }

        rua_a, equip_a = _parse_equip_id(equip_a_id)
        rua_b, equip_b = _parse_equip_id(equip_b_id)
        if rua_a is None or equip_a is None or rua_b is None or equip_b is None:
            return {"success": False, "error": f"IDs inválidos: {equip_a_id} / {equip_b_id}"}

        all_data = [list(row) for row in plano_ws.iter_rows(min_row=2, values_only=True)]
        rows_a: list[dict[str, Any]] = []
        rows_b: list[dict[str, Any]] = []

        for idx, row in enumerate(all_data):
            row_rua = parse_number(row[rua_idx]) if rua_idx < len(row) else None
            row_equip = parse_number(row[equip_idx]) if equip_idx < len(row) else None
            if row_rua is None or row_equip is None:
                continue
            if int(row_rua) == rua_a and int(row_equip) == equip_a:
                rows_a.append({"index": idx, "row": row})
            elif int(row_rua) == rua_b and int(row_equip) == equip_b:
                rows_b.append({"index": idx, "row": row})

        if not rows_a or not rows_b:
            return {
                "success": False,
                "error": f"Equipamento não encontrado. A={len(rows_a)} bins, B={len(rows_b)} bins.",
            }

        if len(rows_a) != len(rows_b):
            return {
                "success": False,
                "error": f"Tamanhos incompatíveis: Equipamento A tem {len(rows_a)} bins, B tem {len(rows_b)} bins.",
            }

        tipo_a = normalize_string(rows_a[0]["row"][tipo_idx] if tipo_idx < len(rows_a[0]["row"]) else "")
        tipo_b = normalize_string(rows_b[0]["row"][tipo_idx] if tipo_idx < len(rows_b[0]["row"]) else "")
        if tipo_a != tipo_b:
            return {"success": False, "error": f"Tipos incompatíveis: A='{tipo_a}', B='{tipo_b}'."}

        updates: dict[int, list[Any]] = {}
        products_to_log: list[dict[str, Any]] = []

        for i in range(len(rows_a)):
            slot_a = rows_a[i]
            slot_b = rows_b[i]

            info_a = _extract_product_info_from_row(slot_a["row"], headers)
            info_b = _extract_product_info_from_row(slot_b["row"], headers)

            new_row_a = [
                _build_new_row_value(headers[col], col, info_b, slot_a["row"], headers)
                for col in range(len(headers))
            ]
            new_row_b = [
                _build_new_row_value(headers[col], col, info_a, slot_b["row"], headers)
                for col in range(len(headers))
            ]

            updates[slot_a["index"]] = new_row_a
            updates[slot_b["index"]] = new_row_b

            if normalize_string(info_a.get("product_code")) not in {"", "Vazio"}:
                products_to_log.append(
                    {
                        "product_code": info_a.get("product_code"),
                        "product_name": normalize_string(info_a.get("product_name")),
                        "from": equip_a_id,
                        "to": equip_b_id,
                    }
                )
            if normalize_string(info_b.get("product_code")) not in {"", "Vazio"}:
                products_to_log.append(
                    {
                        "product_code": info_b.get("product_code"),
                        "product_name": normalize_string(info_b.get("product_name")),
                        "from": equip_b_id,
                        "to": equip_a_id,
                    }
                )

        for idx, new_row in updates.items():
            row_num = idx + 2
            for col_index, value in enumerate(new_row, start=1):
                plano_ws.cell(row=row_num, column=col_index, value=value)

        if log_ws and products_to_log:
            data, hora = _get_log_datetime()
            entries = [
                {
                    "product_code": p["product_code"],
                    "product_name": p.get("product_name") or "",
                    "location_id_anterior": p["from"],
                    "location_id_novo": p["to"],
                    "data": data,
                    "hora": hora,
                    "motivo": "MANUAL-EQUIP-SWAP",
                    "usuario": user,
                }
                for p in products_to_log
            ]
            _append_logs(log_ws, entries)

        wb.save(path)

    return {"success": True, "message": f"Troca de {len(rows_a)} escaninhos concluída."}
