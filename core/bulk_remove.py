from __future__ import annotations

import re
import threading
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .utils import normalize_string

SHEET_BASE_PRODUTOS = "Base_Produtos"
SHEET_PLANO_FINAL = "Plano_Enderecamento_Final"

_lock = threading.Lock()


def _normalize_text(value: Any) -> str:
    text = normalize_string(value).lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.strip()


def _normalize_header(value: Any) -> str:
    text = _normalize_text(value)
    text = text.replace(" ", "_")
    text = re.sub(r"[^a-z0-9_]", "", text)
    return text


def _find_header_index(headers: list[str], candidates: list[str]) -> int:
    normalized = [_normalize_header(h) for h in headers]
    for key in candidates:
        key_norm = _normalize_header(key)
        if key_norm in normalized:
            return normalized.index(key_norm)
    return -1


def _matches_categoria(categoria: str, filtro: str) -> bool:
    if filtro == "prateleira":
        return categoria == "seco" or "prateleira" in categoria
    if filtro == "geladeira":
        return categoria == "refrigerado" or "geladeira" in categoria
    if filtro == "freezer":
        return categoria == "congelado" or "freezer" in categoria
    return False


def _collect_product_codes_xlsx(path: Path, filter_key: str) -> dict[str, Any]:
    normalized_filter = _normalize_text(filter_key)
    valid_filters = {"quimicos", "perfumaria", "tudo_armz", "prateleira", "geladeira", "freezer"}
    if normalized_filter not in valid_filters:
        return {"success": False, "error": "Filtro inválido para remoção total."}

    with _lock:
        wb = load_workbook(path)

        if SHEET_BASE_PRODUTOS not in wb.sheetnames:
            return {"success": False, "error": "Aba Base_Produtos não encontrada"}

        base_ws = wb[SHEET_BASE_PRODUTOS]

        headers = [cell.value for cell in base_ws[1]]
        product_code_idx = _find_header_index(
            headers,
            ["product_code", "codigo_sku", "codigo", "cod_produto", "codigo_produto", "sku"],
        )
        grupo_idx = _find_header_index(headers, ["grupo", "grupo_alocado", "grupo_produto"])
        categoria_idx = _find_header_index(
            headers,
            ["categoria_armazenagem", "categoria_armz", "cat_armz", "categoria"],
        )

        if product_code_idx == -1:
            return {"success": False, "error": "Coluna product_code não encontrada na Base_Produtos."}
        if normalized_filter in {"quimicos", "perfumaria"} and grupo_idx == -1:
            return {"success": False, "error": "Coluna grupo não encontrada na Base_Produtos."}
        if normalized_filter in {"tudo_armz", "prateleira", "geladeira", "freezer"} and categoria_idx == -1:
            return {"success": False, "error": "Coluna categoria_armazenagem não encontrada na Base_Produtos."}

        product_codes_to_remove: set[str] = set()

        for row_num in range(2, base_ws.max_row + 1):
            product_code = normalize_string(base_ws.cell(row=row_num, column=product_code_idx + 1).value)
            if not product_code or product_code == "Vazio":
                continue

            match = False
            if normalized_filter == "quimicos":
                grupo = _normalize_text(base_ws.cell(row=row_num, column=grupo_idx + 1).value)
                match = grupo in {"quimico", "quimicos"}
            elif normalized_filter == "perfumaria":
                grupo = _normalize_text(base_ws.cell(row=row_num, column=grupo_idx + 1).value)
                match = grupo == "perfumaria"
            else:
                categoria = _normalize_text(base_ws.cell(row=row_num, column=categoria_idx + 1).value)
                if normalized_filter == "tudo_armz":
                    match = (
                        _matches_categoria(categoria, "prateleira")
                        or _matches_categoria(categoria, "geladeira")
                        or _matches_categoria(categoria, "freezer")
                    )
                else:
                    match = _matches_categoria(categoria, normalized_filter)

            if match:
                product_codes_to_remove.add(product_code)

    return {
        "success": True,
        "product_codes": product_codes_to_remove,
    }


def preview_remove_all_products_by_filter_xlsx(path: Path, filter_key: str) -> dict[str, Any]:
    result = _collect_product_codes_xlsx(path, filter_key)
    if not result.get("success"):
        return result

    product_codes = result.get("product_codes") or set()
    if not product_codes:
        return {"success": True, "sku_count": 0, "plano_count": 0}

    with _lock:
        wb = load_workbook(path)
        if SHEET_PLANO_FINAL not in wb.sheetnames:
            return {"success": False, "error": "Aba Plano_Enderecamento_Final não encontrada"}

        plano_ws = wb[SHEET_PLANO_FINAL]
        plano_headers = [cell.value for cell in plano_ws[1]]
        plano_product_idx = _find_header_index(
            plano_headers,
            ["product_code", "produto_alocado_code", "codigo_sku", "codigo_produto", "cod_produto"],
        )
        if plano_product_idx == -1:
            return {"success": False, "error": "Coluna product_code não encontrada no Plano_Enderecamento_Final."}

        plano_count = 0
        sku_set: set[str] = set()
        for row_num in range(2, plano_ws.max_row + 1):
            code = normalize_string(plano_ws.cell(row=row_num, column=plano_product_idx + 1).value)
            if code and code in product_codes:
                plano_count += 1
                sku_set.add(code)

    return {"success": True, "sku_count": len(sku_set), "plano_count": plano_count}


def remove_all_products_by_filter_xlsx(path: Path, filter_key: str) -> dict[str, Any]:
    result = _collect_product_codes_xlsx(path, filter_key)
    if not result.get("success"):
        return result

    product_codes = result.get("product_codes") or set()
    if not product_codes:
        return {"success": True, "plano_updated": 0}

    from .moves import _update_row

    with _lock:
        wb = load_workbook(path)
        if SHEET_PLANO_FINAL not in wb.sheetnames:
            return {"success": False, "error": "Aba Plano_Enderecamento_Final não encontrada"}

        plano_ws = wb[SHEET_PLANO_FINAL]
        headers = [normalize_string(cell.value) for cell in plano_ws[1]]
        product_idx = _find_header_index(
            headers,
            ["product_code", "produto_alocado_code", "codigo_sku", "codigo_produto", "cod_produto"],
        )
        if product_idx == -1:
            return {"success": False, "error": "Coluna product_code não encontrada no Plano_Enderecamento_Final."}

        plano_updated = 0
        for row_num in range(2, plano_ws.max_row + 1):
            code = normalize_string(plano_ws.cell(row=row_num, column=product_idx + 1).value)
            if code and code in product_codes:
                original_row = [cell.value for cell in plano_ws[row_num]]
                _update_row(plano_ws, row_num, headers, None, original_row)
                plano_updated += 1

        wb.save(path)

    return {"success": True, "plano_updated": plano_updated}
