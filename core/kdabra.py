from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .data_prep import load_sheet_safe
from .utils import normalize_string

SHEET_PLANO_FINAL = "Plano_Enderecamento_Final"


def generate_kdabra_sheet(path: Path) -> dict[str, Any]:
    plano_data = load_sheet_safe(path, SHEET_PLANO_FINAL)
    wb = load_workbook(path)
    sheet_name = "KDABTA reenderecar"

    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)

    headers = ["cod_produto", "galpao", "rua", "estante", "escaninho"]
    ws.append(headers)

    rows = []
    for row in plano_data:
        product_code = normalize_string(row.get("product_code"))
        if not product_code or product_code == "Vazio":
            continue
        location_id = normalize_string(row.get("location_id"))
        if not location_id:
            continue
        parts = location_id.split("-")
        if len(parts) < 4:
            continue
        galpao = parts[0]
        rua = parts[1]
        estante_raw = parts[2]
        escaninho = "-".join(parts[3:])
        try:
            estante_num = int(estante_raw)
            estante = str(estante_num).zfill(3)
        except ValueError:
            estante = normalize_string(estante_raw).zfill(3)

        rows.append([product_code, galpao, rua, estante, escaninho])

    for row in rows:
        ws.append(row)

    wb.save(path)

    return {"success": True, "url": "/api/download", "sheetName": sheet_name}


def generate_kdabra_enderecar_sheet(path: Path) -> dict[str, Any]:
    plano_data = load_sheet_safe(path, SHEET_PLANO_FINAL)
    wb = load_workbook(path)
    sheet_name = "kdabra enderecar"

    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)

    headers = ["galpao", "rua", "estante", "escaninho", "ordem"]
    ws.append(headers)

    unique_locations: dict[str, dict[str, Any]] = {}
    for row in plano_data:
        location_id = normalize_string(row.get("location_id"))
        if not location_id:
            continue
        if location_id in unique_locations:
            continue
        parts = location_id.split("-")
        if len(parts) < 4:
            continue
        galpao = parts[0]
        rua = parts[1]
        estante = parts[2]
        escaninho = "-".join(parts[3:])
        nivel = "0"
        posicao = ""
        nivel_match = __import__("re").match(r"^(\d+)", escaninho)
        pos_match = __import__("re").search(r"([A-Z]+)$", escaninho)
        if nivel_match:
            nivel = nivel_match.group(1)
        if pos_match:
            posicao = pos_match.group(1)

        try:
            rua_num = int(rua.replace("R", "").replace("r", ""))
        except ValueError:
            rua_num = 0
        try:
            estante_num = int(estante)
        except ValueError:
            estante_num = 0

        unique_locations[location_id] = {
            "galpao": galpao,
            "rua": rua,
            "estante": estante,
            "escaninho": escaninho,
            "ruaNum": rua_num,
            "estanteNum": estante_num,
            "nivel": nivel,
            "posicao": posicao,
        }

    all_locations = list(unique_locations.values())

    def sort_key(loc: dict[str, Any]) -> tuple[int, int, int, str]:
        nivel = int(loc["nivel"]) if str(loc["nivel"]).isdigit() else 0
        return (loc["ruaNum"], loc["estanteNum"], -nivel, loc["posicao"])

    all_locations.sort(key=sort_key)
    for idx, loc in enumerate(all_locations, start=1):
        estante_formatted = str(loc["estanteNum"]).zfill(3)
        ws.append([loc["galpao"], loc["rua"], estante_formatted, loc["escaninho"], idx])

    wb.save(path)

    return {"success": True, "url": "/api/download", "sheetName": sheet_name}
