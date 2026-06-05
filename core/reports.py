from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .data_prep import (
    SHEET_BASE_PRODUTOS,
    build_base_produtos_map,
    build_dic_cat_map,
    load_limits,
    load_sheet_safe,
)
from .utils import parse_bool_flag, normalize_string

SHEET_PLANO_FINAL = "Plano_Enderecamento_Final"


def _build_report_dataset(
    source: Any, abas: list[str], colunas: list[str]
) -> tuple[dict[str, list[list[Any]]], list[str], dict[str, Any]]:
    base_data = load_sheet_safe(source, SHEET_BASE_PRODUTOS)
    plano_data = load_sheet_safe(source, SHEET_PLANO_FINAL)
    dic_cat_map = build_dic_cat_map(load_sheet_safe(source, "Dicionario_Categorias"))
    limite_altura, _, limite_altura_baixo = load_limits(load_sheet_safe(source, "Configuracoes_Operacionais"))
    base_map = build_base_produtos_map(base_data, dic_cat_map, limite_altura, limite_altura_baixo)

    aba_map = {
        "refrigerados": {"name": "Refrigerados", "types": ["geladeira", "geladeira_alta", "geladeira_americana"], "filter": None},
        "congelados": {"name": "Congelados", "types": ["freezer"], "filter": None},
        "secos": {"name": "Secos", "types": ["prateleira", "prateleira lateral"], "filter": None},
        "nao-alocados": {"name": "Nao Alocados", "types": [], "filter": None},
        "altos": {"name": "Produtos Altos", "types": None, "filter": {"is_alto": True}},
        "pesados": {"name": "Produtos Pesados", "types": None, "filter": {"is_pesado": True}},
        "pequenos": {"name": "Produtos Pequenos", "types": None, "filter": {"is_pequeno": True}},
        "frageis": {"name": "Produtos Frágeis", "types": None, "filter": {"is_fragil": "SIM"}},
        "degelo-nao": {"name": "Degelo = NÃO", "types": None, "filter": {"degelo": "NAO"}},
        "curva-a": {"name": "Curva A", "types": None, "filter": {"curva": "A"}},
        "curva-b": {"name": "Curva B", "types": None, "filter": {"curva": "B"}},
        "curva-c": {"name": "Curva C", "types": None, "filter": {"curva": "C"}},
        "curva-d": {"name": "Curva D", "types": None, "filter": {"curva": "D"}},
        "curva-e": {"name": "Curva E", "types": None, "filter": {"curva": "E"}},
        "curva-cd": {"name": "Curva C+D", "types": None, "filter": {"curva": ["C", "D"]}},
        "sem-curva": {"name": "Sem Curva", "types": None, "filter": {"curva": None}},
        "grupo-quimicos": {"name": "Grupo: Químicos", "types": None, "filter": {"grupo": "quimicos"}},
        "grupo-perfumaria": {"name": "Grupo: Perfumaria", "types": None, "filter": {"grupo": "perfumaria"}},
        "grupo-alimento": {"name": "Grupo: Alimento", "types": None, "filter": {"grupo": "alimento"}},
        "grupo-flvs": {"name": "Grupo: FLVs", "types": None, "filter": {"grupo": "flvs"}},
        "grupo-bebidas": {"name": "Grupo: Bebidas", "types": None, "filter": {"grupo": "bebidas"}},
        "grupo-neutro": {"name": "Grupo: Neutro", "types": None, "filter": {"grupo": "neutro"}},
    }

    coluna_map = {
        "codigo_sku": 0,
        "descricao": 1,
        "endereco": 2,
        "tipo_equipamento": 3,
        "categoria_armazenagem": 4,
        "curva": 5,
        "grupo": 6,
        "fabricante": 7,
        "quantidade": 8,
        "alto": 9,
        "pesado": 10,
        "pequeno": 11,
        "fragil": 12,
        "degelo": 13,
        "altura": 14,
        "peso": 15,
        "volume": 16,
        "subcategoria": 17,
        "vendas": 18,
    }

    headers_map = {
        "codigo_sku": "codigo_sku",
        "descricao": "descricao",
        "endereco": "endereco",
        "tipo_equipamento": "tipo_equipamento",
        "categoria_armazenagem": "categoria_armazenagem",
        "curva": "curva",
        "grupo": "grupo",
        "fabricante": "fabricante",
        "quantidade": "quantidade",
        "alto": "alto",
        "pesado": "pesado",
        "pequeno": "pequeno",
        "fragil": "fragil",
        "degelo": "degelo",
        "altura": "altura_cm",
        "peso": "peso_kg_unitario",
        "volume": "vol_l_unitario",
        "subcategoria": "subcategoria",
        "vendas": "venda_total",
    }

    headers = [headers_map.get(col, col) for col in colunas]

    alocated_codes = set()
    report_data: dict[str, list[list[Any]]] = {}
    for aba in abas:
        if aba in aba_map:
            report_data[aba_map[aba]["name"]] = []

    for row in plano_data:
        pcode = normalize_string(row.get("product_code"))
        if not pcode or pcode == "Vazio":
            continue
        alocated_codes.add(pcode)
        product_info = base_map.get(pcode, {})
        tipo_equip = normalize_string(row.get("tipo_equipamento_final") or row.get("tipo_equipamento") or "N/A")

        altura = product_info.get("altura_cm") or row.get("altura_cm")
        peso = product_info.get("peso_kg_unitario") or row.get("peso_kg_unitario")
        is_alto = parse_bool_flag(product_info.get("is_alto"))
        is_pesado = parse_bool_flag(product_info.get("is_pesado"))
        is_pequeno = parse_bool_flag(product_info.get("is_pequeno"))
        is_fragil = product_info.get("is_fragil") or row.get("is_fragil") or "N/A"
        degelo = product_info.get("degelo") or row.get("degelo") or "N/A"
        vol_l = product_info.get("vol_l_unitario") or row.get("vol_l_unitario") or product_info.get("vol_L_unitario") or row.get("vol_L_unitario")
        venda_total = product_info.get("venda_total") or row.get("venda_total")

        full_row = [
            pcode,
            product_info.get("product_name") or row.get("product_name") or "Produto sem nome",
            row.get("location_id") or "N/A",
            tipo_equip,
            product_info.get("categoria_armazenagem") or row.get("categoria_armazenagem") or "N/A",
            product_info.get("curva") or row.get("curva") or "N/A",
            product_info.get("grupo") or row.get("grupo") or "N/A",
            product_info.get("nm_fabricante") or row.get("nm_fabricante") or "N/A",
            row.get("quantidade") or product_info.get("quantidade") or "N/A",
            "SIM" if is_alto else "NÃO",
            "SIM" if is_pesado else "NÃO",
            "SIM" if is_pequeno else "NÃO",
            is_fragil or "N/A",
            degelo or "N/A",
            altura if altura is not None else "N/A",
            peso if peso is not None else "N/A",
            vol_l if vol_l is not None else "N/A",
            product_info.get("subcategoria") or row.get("subcategoria") or "N/A",
            venda_total if venda_total is not None else "N/A",
        ]

        filtered_row = [full_row[coluna_map[col]] for col in colunas]

        for aba in abas:
            config = aba_map.get(aba)
            if not config:
                continue

            should_include = False
            if config["filter"]:
                filt = config["filter"]
                if "is_alto" in filt and filt["is_alto"] != is_alto:
                    continue
                if "is_pesado" in filt and filt["is_pesado"] != is_pesado:
                    continue
                if "is_pequeno" in filt and filt["is_pequeno"] != is_pequeno:
                    continue
                if "is_fragil" in filt:
                    fragil_val = normalize_string(is_fragil).upper()
                    if filt["is_fragil"] == "SIM" and fragil_val != "SIM":
                        continue
                if "degelo" in filt:
                    degelo_val = normalize_string(degelo).upper()
                    if filt["degelo"] == "NAO" and degelo_val != "NAO":
                        continue
                if "curva" in filt:
                    curva_val = normalize_string(product_info.get("curva") or row.get("curva")).upper()
                    if isinstance(filt["curva"], list):
                        if curva_val not in [c.upper() for c in filt["curva"]]:
                            continue
                    elif filt["curva"] is None:
                        if curva_val not in {"", "N/A"}:
                            continue
                    else:
                        if curva_val != str(filt["curva"]).upper():
                            continue
                if "grupo" in filt:
                    grupo_val = normalize_string(product_info.get("grupo") or row.get("grupo")).lower()
                    if grupo_val != str(filt["grupo"]).lower():
                        continue
                should_include = True
            elif config["types"]:
                if tipo_equip in config["types"]:
                    should_include = True
            elif aba == "nao-alocados":
                continue

            if should_include:
                report_data[config["name"]].append(filtered_row)
                break

    processed_non_allocated = set()
    for pcode, product_info in base_map.items():
        is_allocated = pcode in alocated_codes
        altura = product_info.get("altura_cm")
        peso = product_info.get("peso_kg_unitario")
        is_alto = parse_bool_flag(product_info.get("is_alto"))
        is_pesado = parse_bool_flag(product_info.get("is_pesado"))
        is_pequeno = parse_bool_flag(product_info.get("is_pequeno"))
        is_fragil = product_info.get("is_fragil") or "N/A"
        degelo = product_info.get("degelo") or "N/A"
        vol_l = product_info.get("vol_l_unitario") or product_info.get("vol_L_unitario")
        venda_total = product_info.get("venda_total")
        curva = product_info.get("curva") or ""
        grupo = product_info.get("grupo") or ""

        full_row = [
            pcode,
            product_info.get("product_name") or "Produto sem nome",
            "N/A",
            "N/A",
            product_info.get("categoria_armazenagem") or "N/A",
            curva or "N/A",
            grupo or "N/A",
            product_info.get("nm_fabricante") or "N/A",
            product_info.get("quantidade") or "N/A",
            "SIM" if is_alto else "NÃO",
            "SIM" if is_pesado else "NÃO",
            "SIM" if is_pequeno else "NÃO",
            is_fragil or "N/A",
            degelo or "N/A",
            altura if altura is not None else "N/A",
            peso if peso is not None else "N/A",
            vol_l if vol_l is not None else "N/A",
            product_info.get("subcategoria") or "N/A",
            venda_total if venda_total is not None else "N/A",
        ]
        filtered_row = [full_row[coluna_map[col]] for col in colunas]

        for aba in abas:
            config = aba_map.get(aba)
            if not config:
                continue

            should_include = False
            if aba == "nao-alocados" and not is_allocated:
                should_include = True
            elif config["filter"]:
                filt = config["filter"]
                if "is_alto" in filt and filt["is_alto"] != is_alto:
                    continue
                if "is_pesado" in filt and filt["is_pesado"] != is_pesado:
                    continue
                if "is_pequeno" in filt and filt["is_pequeno"] != is_pequeno:
                    continue
                if "is_fragil" in filt:
                    fragil_val = normalize_string(is_fragil).upper()
                    if filt["is_fragil"] == "SIM" and fragil_val != "SIM":
                        continue
                if "degelo" in filt:
                    degelo_val = normalize_string(degelo).upper()
                    if filt["degelo"] == "NAO" and degelo_val != "NAO":
                        continue
                if "curva" in filt:
                    curva_val = normalize_string(curva).upper()
                    if isinstance(filt["curva"], list):
                        if curva_val not in [c.upper() for c in filt["curva"]]:
                            continue
                    elif filt["curva"] is None:
                        if curva_val not in {"", "N/A"}:
                            continue
                    else:
                        if curva_val != str(filt["curva"]).upper():
                            continue
                if "grupo" in filt:
                    grupo_val = normalize_string(grupo).lower()
                    if grupo_val != str(filt["grupo"]).lower():
                        continue
                should_include = True

            if should_include and f"{pcode}-{aba}" not in processed_non_allocated:
                report_data[config["name"]].append(filtered_row)
                processed_non_allocated.add(f"{pcode}-{aba}")
                break

    return report_data, headers, aba_map


def generate_sku_report_custom(
    source: Any, destination: str, abas: list[str], colunas: list[str]
) -> dict[str, Any]:
    report_data, headers, aba_map = _build_report_dataset(source, abas, colunas)

    if destination == "new":
        output = io.StringIO()
        writer = csv.writer(output)
        first = True
        for aba in abas:
            config = aba_map.get(aba)
            if not config:
                continue
            sheet_name = f"Relatório SKUs - {config['name']}"
            if not first:
                output.write("\n\n")
            first = False
            output.write(f"=== {sheet_name} ===\n")
            writer.writerow(headers)
            for row in report_data.get(config["name"], []):
                writer.writerow(row)
        return {
            "success": True,
            "csvContent": output.getvalue(),
            "filename": "Relatorio_SKUs.csv",
        }

    if hasattr(source, "read_sheet"):
        return {"success": False, "error": "Destino 'same' requer planilha local."}

    wb = load_workbook(Path(source))
    created = []
    for aba in abas:
        config = aba_map.get(aba)
        if not config:
            continue
        sheet_name = f"Relatório SKUs - {config['name']}"
        if sheet_name in wb.sheetnames:
            idx = wb.sheetnames.index(sheet_name)
            wb.remove(wb[sheet_name])
            ws = wb.create_sheet(sheet_name, idx)
        else:
            ws = wb.create_sheet(sheet_name)
        created.append(sheet_name)

        ws.append(headers)
        for row in report_data.get(config["name"], []):
            ws.append(row)

    wb.save(path)

    return {
        "success": True,
        "url": "/api/download",
        "sheetName": created[0] if created else None,
    }
