from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .utils import normalize_string

VERSION_DIR = Path(__file__).resolve().parents[1] / ".versions"
VERSION_PREFIX = "plano"


def _sanitize_name(name: str) -> str:
    text = normalize_string(name)
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text or "sem_nome"


def _timestamp() -> str:
    return datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d_%H%M%S")


def _build_version_filename(name: str) -> str:
    safe = _sanitize_name(name)
    return f"{VERSION_PREFIX}_{_timestamp()}__{safe}.xlsx"


def save_plano_version_xlsx(path: Path, name: str) -> dict[str, Any]:
    if not path.exists():
        return {"success": False, "error": "Arquivo base não encontrado."}

    VERSION_DIR.mkdir(parents=True, exist_ok=True)
    filename = _build_version_filename(name)
    dest = VERSION_DIR / filename
    shutil.copy2(path, dest)

    return {"success": True, "version_id": filename, "label": filename}


def list_plano_versions_xlsx(_: Path) -> dict[str, Any]:
    if not VERSION_DIR.exists():
        return {"success": True, "versions": []}

    versions = []
    for file in VERSION_DIR.glob("*.xlsx"):
        label = file.stem.replace(f"{VERSION_PREFIX}_", "").replace("__", " ")
        versions.append(
            {
                "version_id": file.name,
                "label": label or file.name,
                "timestamp": file.stat().st_mtime,
            }
        )

    versions.sort(key=lambda v: v.get("timestamp", 0), reverse=True)
    return {"success": True, "versions": versions}


def restore_plano_version_xlsx(path: Path, version_id: str) -> dict[str, Any]:
    if not path.exists():
        return {"success": False, "error": "Arquivo base não encontrado."}
    version_path = VERSION_DIR / version_id
    if not version_path.exists():
        return {"success": False, "error": "Versão não encontrada."}

    src_wb = load_workbook(version_path)
    dst_wb = load_workbook(path)

    if "Plano_Enderecamento_Final" not in src_wb.sheetnames:
        return {"success": False, "error": "Aba Plano_Enderecamento_Final não encontrada na versão."}
    if "Plano_Enderecamento_Final" not in dst_wb.sheetnames:
        return {"success": False, "error": "Aba Plano_Enderecamento_Final não encontrada no arquivo atual."}

    src_ws = src_wb["Plano_Enderecamento_Final"]
    dst_ws = dst_wb["Plano_Enderecamento_Final"]

    # Limpar conteúdo atual
    if dst_ws.max_row > 0 and dst_ws.max_column > 0:
        for row in dst_ws.iter_rows(min_row=1, max_row=dst_ws.max_row, max_col=dst_ws.max_column):
            for cell in row:
                cell.value = None

    # Copiar valores
    max_row = src_ws.max_row
    max_col = src_ws.max_column
    for row in src_ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

    dst_wb.save(path)
    return {"success": True, "rows": max_row, "cols": max_col}


def delete_plano_version_xlsx(version_id: str) -> dict[str, Any]:
    version_path = VERSION_DIR / version_id
    if not version_path.exists():
        return {"success": False, "error": "Versão não encontrada."}
    version_path.unlink()
    return {"success": True}
